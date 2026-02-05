"""
Excel export helper functions for mess management system
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


def create_excel_response(filename):
    """Create an HTTP response for Excel file download"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_header_row(worksheet, row_num=1):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in worksheet[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_adjust_column_width(worksheet):
    """Auto-adjust column widths based on content"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_payments_to_excel(payments, month_year):
    """
    Export payments to Excel file
    
    Args:
        payments: QuerySet of Payment objects
        month_year: Month/year string for filename
    
    Returns:
        HttpResponse with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payments {month_year}"
    
    # Headers
    headers = ['User', 'Month', 'Amount (₹)', 'Status', 'Transaction ID', 'Paid Date']
    ws.append(headers)
    style_header_row(ws)
    
    # Data rows
    for payment in payments:
        ws.append([
            payment.user.get_full_name(),
            payment.month_year,
            float(payment.amount),
            payment.get_status_display(),
            payment.transaction_id or 'N/A',
            payment.paid_date.strftime('%Y-%m-%d %H:%M') if payment.paid_date else 'N/A'
        ])
    
    # Add totals row
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value='Total Collected:')
    ws.cell(row=total_row, column=3, value=f'=SUMIF(D2:D{ws.max_row-2},"Paid",C2:C{ws.max_row-2})')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    
    auto_adjust_column_width(ws)
    
    # Create response
    response = create_excel_response(f'payments_{month_year}.xlsx')
    wb.save(response)
    return response


def export_groceries_to_excel(groceries, month_year):
    """Export groceries to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Groceries {month_year}"
    
    # Headers
    headers = ['Item Name', 'Category', 'Quantity', 'Price (₹)', 'Purchase Date']
    ws.append(headers)
    style_header_row(ws)
    
    # Data rows
    for grocery in groceries:
        ws.append([
            grocery.item_name,
            grocery.get_category_display(),
            grocery.quantity,
            float(grocery.price),
            grocery.purchase_date.strftime('%Y-%m-%d')
        ])
    
    # Add totals row
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value='Total Grocery Expenses:')
    ws.cell(row=total_row, column=4, value=f'=SUM(D2:D{ws.max_row-2})')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    auto_adjust_column_width(ws)
    
    response = create_excel_response(f'groceries_{month_year}.xlsx')
    wb.save(response)
    return response


def export_monthly_report_to_excel(month_year, payments, groceries, fixed_expense):
    """Export comprehensive monthly report to Excel"""
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = f'Mess Management Report - {month_year}'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    # Payment Summary
    ws_summary['A3'] = 'Payment Summary'
    ws_summary['A3'].font = Font(bold=True, size=14)
    
    total_collected = sum(p.amount for p in payments if p.status == 'paid')
    pending_count = sum(1 for p in payments if p.status == 'pending')
    
    ws_summary.append(['', ''])  # Empty row
    ws_summary.append(['Total Collected:', f'₹{total_collected:.2f}'])
    ws_summary.append(['Pending Payments:', pending_count])
    ws_summary.append(['Total Users:', payments.count()])
    
    # Expense Summary
    ws_summary['A9'] = 'Expense Summary'
    ws_summary['A9'].font = Font(bold=True, size=14)
    
    total_grocery = sum(g.price for g in groceries)
    total_fixed = fixed_expense.total_fixed_expense if fixed_expense else 0
    
    ws_summary.append(['', ''])
    ws_summary.append(['Grocery Expenses:', f'₹{total_grocery:.2f}'])
    ws_summary.append(['Fixed Expenses:', f'₹{total_fixed:.2f}'])
    ws_summary.append(['Total Expenses:', f'₹{total_grocery + total_fixed:.2f}'])
    
    auto_adjust_column_width(ws_summary)
    
    # Payments Sheet
    ws_payments = wb.create_sheet("Payments")
    headers = ['User', 'Amount (₹)', 'Status', 'Transaction ID']
    ws_payments.append(headers)
    style_header_row(ws_payments)
    
    for payment in payments:
        ws_payments.append([
            payment.user.get_full_name(),
            float(payment.amount),
            payment.get_status_display(),
            payment.transaction_id or 'N/A'
        ])
    auto_adjust_column_width(ws_payments)
    
    # Groceries Sheet
    if groceries:
        ws_groceries = wb.create_sheet("Groceries")
        headers = ['Item', 'Category', 'Quantity', 'Price (₹)']
        ws_groceries.append(headers)
        style_header_row(ws_groceries)
        
        for grocery in groceries:
            ws_groceries.append([
                grocery.item_name,
                grocery.get_category_display(),
                grocery.quantity,
                float(grocery.price)
            ])
        auto_adjust_column_width(ws_groceries)
    
    response = create_excel_response(f'monthly_report_{month_year}.xlsx')
    wb.save(response)
    return response
